# importing openpyxl module
import openpyxl
from backend.Database import Database
# Give the location of the file
path = "G:\\thaihmong_translation\\test\\tnc_free.xlsx"

# workbook object is created
wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active
m_row = sheet_obj.max_row

# Loop will print all values
# of first column
total = []
for i in range(53786,53893 ):
    try:
        word = sheet_obj.cell(row=i, column=1)
        fre = sheet_obj.cell(row=i, column=2)

        word = str(word.value).strip()
        fre = str(fre.value).strip()

        # print(word," ",fre)

        db = Database()
        data = db.search__(word)
        # print("data :",data)

        if (int(data) > 0):
            total.append([word, fre])
        ll = len(total)
        print(ll)
        if(ll>10100):
            break
    except Exception as e:
        print(e)

print("\n--------------------------------\n")

for i in total:
    print(i[0],"\t",i[1])


